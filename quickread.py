class Define():
    allf = []
    class Document():
        def __init__(self,adress,content,type):
            self.adress = adress
            self.content = content
            self.type = type
            self.usage = 0

    def __init__(self,read):
        self.read = read
        self.adi = {}
        Define.allf.append(self)
    def open(self,adress):
        if adress in self.adi:
            obj = self.adi[adress]
        else:
            f = self.read
            content = error_proof_opener(f,adress) #return False if error while opening
            obj = self.Document(adress,content,self.read)
            self.adi[adress] = obj

        obj.usage += 1
        return obj.content

def error_proof_opener(f,adress):
    opened = False
    errorshowed = False
    loop = 0
    while opened == False:
        message = None
        try:
            content = f(adress)
        except PermissionError:
            message = "Please, close the app that's using the file"
            time.sleep(0.2)
            loop = 0
        except FileNotFoundError:
            message = "Trying automatic correction of a temporary file"
            adress = adress.replace("~$","")
        else:
            opened = True
        finally:
            if message != None:
                print(message)
            loop += 1

        if loop >= 5:
            print("Can't open the file, does it even exist ?")
            return False
    return content

def showusages(func=None):
    from tabulate import tabulate
    global_data = []
    if func == None:
        func = Define.allf
    if callable(func):
        func = [func]

    for f in func:
        local_data = []
        for doc in f.adi.values():
            global_data.append((doc.type,doc.adress,doc.usage))
    print(tabulate(global_data,headers=['Function','Adress','Open x times']))

def UpgradeOpenpyxl():

    import openpyxl
    import pandas
    from itertools import islice

    def xlsxsheet_to_df(sheet):

        data = sheet.values
        cols = next(data)[1:]
        data = list(data)
        idx = [r[0] for r in data]
        data = (islice(r, 1, None) for r in data)
        df = pd.DataFrame(data, index=idx, columns=cols)
        return df

    def xlsx_to_df(workbook):
        return [df for df in xlsxsheet_to_df(sheet) for sheet in worbook]

    def df_to_xlsx(df):
        wb = openpyxl.Workbook()
        ws = wb.active

        for r in dataframe_to_rows(df, index=True, header=True):
            ws.append(r)

        for cell in ws['A'] + ws[1]:
            cell.style = 'Pandas'

        return wb


    openpyxl.qckopen = Define(openpyxl.open).open
    openpyxl.qckload_worbook = Define(openpyxl.load_workbook).open
    openpyxl.workbook_to_dfs = xlsx_to_df
    openpyxl.sheet_to_df = xlsxsheet_to_df
    openpyxl.df_to_xlsx = df_to_xlsx

def testing(loop=100):
    import time
    import openpyxl
    import tempfile
    UpgradeOpenpyxl()
    L = []
    for n in range(loop):
        file = tempfile.NamedTemporaryFile(suffix=".xlsx",delete = False)
        W = openpyxl.Workbook()
        S = W.active
        S.cell(1,1).value = f"test {n}"
        W.save(file.name)
        L.append(file.name)

    at = time.perf_counter()
    for file_name in L:
        openpyxl.open(file_name)
    bt = time.perf_counter()
    for file_name in L:
        openpyxl.qckopen(file_name) #the first time the loop is being looped, it need to load the file with classic openpyxl.open() that's why I made a second for loop below for pure optimized time perf.
    ct = time.perf_counter()
    for file_name in L:
         openpyxl.qckopen(file_name)
    dt = time.perf_counter()

    print('Loading time without optimisation :',bt-at)
    print('Loading with optimization, and with one real reading :',ct-bt)
    print('Loading time only optimized :',dt-ct)
    print(f"Différence with/without : {round((bt-at)/(dt-ct),1)} x times faster !")



